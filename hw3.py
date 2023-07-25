from constants import *
from datetime import datetime, timedelta
from airflow import DAG
from airflow.operators.bash import BashOperator
from airflow.operators.python import PythonOperator
from airflow.utils.dates import days_ago
from airflow.providers.sqlite.hooks.sqlite import SqliteHook
import requests
import asyncio
import sqlite3 as sql
import time
from aiohttp import ClientSession as cs
from aiohttp import TCPConnector as tcpcn
import json
import logging
import logging.config
from pathlib import Path
import zipfile
import pandas as pd
from normalizer import normal
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

config = Path('/home/kirill/airflow/dags/logger.conf')
logging.config.fileConfig(fname=config, disable_existing_loggers=False)
logger = logging.getLogger('hwLogger')

default_args = {
    'owner': 'kirill',
    'retries': 1,
    'retry_delay': timedelta(seconds=10)
}

def telecoms(**kwargs):
    try:
        sh = SqliteHook(sqlite_conn_id='hw3')
        connection = sh.get_conn()
        cursor = connection.cursor()
        logger.info(f"База данных {db} подключена")
        cursor.execute(tab_telecom)
        connection.commit()
        logger.info("Таблица telecom_companies успешно добавлена в БД")
        with zipfile.ZipFile(egrul_local, 'r') as zip:
            names = zip.namelist()
            for name in names:
                data = []
                with zip.open(name) as file:
                    egrul = pd.read_json(file)
                    for index, row in egrul.iterrows():
                        ogrn_date = ''
                        if 'ДатаОГРН' in row['data'].keys():
                            ogrn_date = row['data']['ДатаОГРН']
                        if 'СвОКВЭД' in row['data'].keys():
                            if 'СвОКВЭДОсн' in row['data']['СвОКВЭД'].keys():
                                okved = row['data']['СвОКВЭД']['СвОКВЭДОсн']['КодОКВЭД']
                                if okved[:2] == '61':
                                    data.append([row['name'], normal(row['name']), row['inn'], row['ogrn'], ogrn_date, okved])
                    try:
                        cursor.executemany(insert_telecom, data)
                        connection.commit()
                        logger.info(f"Данные файла {name} успешно добавлены в таблицу telecom_companies")
                    except sql.Error as error:
                        logger.error(f"Не удалось добавить данные файла {name} в таблицу telecom_companies")
                        logger.error(f"Исключение: {error.__class__}, {error.args}")
        cursor.close()
    except sql.Error as error:
        logger.critical("Не удалось добавить таблицу telecom_companies")
        logger.critical(f"Исключение: {error.__class__}, {error.args}")
    except Exception:
        logger.critical(f"Ошибка чтения архива {egrul_local}")
    finally:                    
        if (connection):
            connection.close()
            logger.info("Соединение закрыто")

async def get_vacancy(url, session):
    async with session.get(url=url) as vacancy_page:
        vacancy_json = await vacancy_page.json()
        return vacancy_json
        
async def main(items, data):
    connector = tcpcn(limit=10)
    async with cs(api_url, connector=connector) as session:
        tasks = []
        for item in items:
            url = f"/{item['url'][18:]}"
            tasks.append(asyncio.create_task(get_vacancy(url, session)))
        vacancies_json = await asyncio.gather(*tasks)
    for vacancy in vacancies_json:
        description = ''
        key_skills = ''
        if (vacancy['description'] != None):
            description = vacancy['description']
        if (vacancy['key_skills'] != None):
            for skill in vacancy['key_skills']:
                key_skills += skill['name'] + ', '
            key_skills = key_skills[:-2]
        data.append([vacancy['employer']['name'], normal(vacancy['employer']['name']), vacancy['name'], description, key_skills])

def vacancies(**kwargs):
    try:
        search_result = requests.get(url, params=url_params)
        if search_result.status_code == 200:
            vacancy_list = search_result.json().get('items')
            data = []
            start = time.time()
            asyncio.run(main(vacancy_list, data))
            logger.info(f"Время выполнения, с: {time.time() - start}")
            sh = SqliteHook(sqlite_conn_id='hw3')
            try:
                connection = sh.get_conn()
                cursor = connection.cursor()
                logger.info(f"База данных {db} подключена")
                cursor.execute(tab_vacancies)
                connection.commit()
                logger.info('Таблица vacancies успешно добавлена в БД')
                cursor.executemany(insert_vacancy, data)
                connection.commit()
                logger.info('Данные успешно добавлены в таблицу vacancies')
                cursor.close()
            except sql.Error as error:
                logger.error('Не удалось вставить данные в таблицу vacancies')
                logger.error("Исключение: ", error.__class__, error.args)
            finally:
                if (connection):
                    connection.close()
                    logger.info("Соединение с закрыто")
        else:
            logger.critical(f"Не удалось загрузить данные поиска: {search_result.status_code}")
    except Exception:
        logger.critical(f"Исключение: {Exception.__class__}, {Exception.args}")

def skills(**kwargs):
    try:
        sh = SqliteHook(sqlite_conn_id='hw3')
        connection = sh.get_conn()
        cursor = connection.cursor()
        logger.info(f"База данных {db} подключена")
        joiner = """
            SELECT vacancies.employer_name, vacancies.key_skills, telecom_companies.name, max(telecom_companies.okved)
            FROM vacancies INNER JOIN telecom_companies
            ON telecom_companies.normal_name LIKE vacancies.normal_name
            GROUP BY vacancies.employer_name;"""
        cursor.execute(joiner)
        skills = []
        for item in cursor.fetchall():
            skills.extend(item[1].replace(' ', '').split(','))
            logger.info(item)
        c = Counter(skills)
        cursor.execute(tab_skills_top)
        connection.commit()
        logger.info('Таблица skills_top успешно добавлена в БД')
        wb = Workbook()
        ws = wb.active
        ws.append(['Количество', 'Скилл'])
        ft = Font(bold=True)
        ws['A1'].font = ft
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['B1'].font = ft
        ws['B1'].alignment = Alignment(horizontal='center')
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        for skill in c.most_common(10):
            cursor.execute(insert_skill, [skill[1], skill[0]])
            ws.append([skill[1], skill[0]])
            logger.info(f"{str(skill[1]).rjust(2)} - {skill[0]}")
        connection.commit()
        logger.info('Данные успешно добавлены в таблицу skills_top')
        cursor.close()
        wb.save('/home/kirill/skills.xlsx')
    except sql.Error as error:
        logger.critical("Не удалось выполнить запрос")
        logger.critical(f"Исключение: {error.__class__}, {error.args}")
    finally:                    
        if (connection):
            connection.close()
            logger.info("Соединение закрыто")

with DAG(
    dag_id='hw3',
    default_args=default_args,
    description='DAG for home work 3',
    start_date=datetime(2023, 7, 22),
    schedule='@daily'
) as dag:
    download_egrul = BashOperator(
        task_id='download_egrul',
        bash_command=f"wget {egrul_web} -O {egrul_local}"
    )
    load_telecoms = PythonOperator(
        task_id='load_telecoms',
        python_callable=telecoms,
    )
    get_vacancies = PythonOperator(
        task_id='get_vacancies',
        python_callable=vacancies,
    )
    top_skills = PythonOperator(
        task_id='top_skills',
        python_callable=skills,
    )
    download_egrul >> load_telecoms >> get_vacancies >> top_skills